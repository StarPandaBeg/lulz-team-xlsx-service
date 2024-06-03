
import openpyxl
from sqlalchemy.exc import DataError
from openpyxl.styles import Border, Side, Font
from openpyxl.styles import PatternFill
from flask import request, jsonify, send_file, after_this_request

import numpy as np
import cv2
import os
from config import *
from appenv import *
import db


@app.route("/export_to_excel/<int:id>", methods=['GET'])
def export_to_excel(id):
    session = db.SessionLocal()
    data = session.query(db.TransactionsWithReceipts).filter_by(
        komandirovka_id=id).all()
    wb = openpyxl.Workbook()
    ws = wb.active

    # create mini table
    credit_summ = 0
    debit_summ = 0

    for row in data:
        res = float(row.amount_in_account_currency)

        if "debit" in row.operation_type.lower():
            debit_summ += res
        elif "credit" in row.operation_type.lower():
            credit_summ += res

    query = "SELECT * FROM komandirovki WHERE id = :id"
    params = {'id': id}

    result = session.execute(db.totext(query), params)

    column_labels = {
        "fio": "ФИО",
        "date_start": "Время начала",
        "date_end": "Время окончания",
    }
    headers = [column_labels.get(col, col) for col in result.keys()]
    headers += ["Приход", "Расход"]
    ws.append(headers)

    for row in result:
        ws.append(tuple(row) + (credit_summ, debit_summ))

    ws.append([])  # insert empty row

    # create big table
    column_labels = {
        "account_number": "Номер счета",
        "transaction_id": "Идентификатор транзакции",
        "operation_type": "Тип операции (пополнение/списание)",
        "operation_category": "Категория операции",
        "status": "Статус",
        "creation_date": "Дата создания операции",
        "authorization_date": "Дата авторизации",
        "transaction_date": "Дата транзакции",
        "original_operation_id": "Идентификатор оригинальной операции",
        "operation_amount_in_currency": "Сумма операции в валюте операции",
        "operation_currency": "Валюта операции",
        "amount_in_account_currency": "Сумма в валюте счета",
        "counterparty": "Контрагент",
        "counterparty_inn": "ИНН контрагента",
        "counterparty_kpp": "КПП контрагента",
        "counterparty_account": "Счет контрагента",
        "counterparty_bank_bik": "БИК банка контрагента",
        "counterparty_bank_corr_account": "Корр. счет банка контрагента",
        "counterparty_bank_name": "Наименование банка контрагента",
        "payment_purpose": "Назначение платежа",
        "payment_number": "Номер платежа",
        "sequence": "Очередность",
        "code": "Код (УИН)",
        "card_number": "Номер карты",
        "mcc": "MCC",
        "place_of_transaction_city": "Место совершения (Город)",
        "place_of_transaction_country": "Место совершения (Страна)",
        "organization_address": "Адрес организации",
        "bank": "Банк",
        "document_creator_status": "Статус составителя расчетного документа",
        "budget_classification_code": "КБК-код бюджетной классификации",
        "oktmo_code": "Код ОКТМО",
        "tax_payment_basis": "Основание налогового платежа",
        "tax_period_customs_code": "Налоговый период/Код таможенного органа",
        "tax_document_number": "Номер налогового документа",
        "tax_document_date": "Дата налогового документа",
        "tax_payment_type": "Тип налогового платежа",
        "receipt_fd": "ФД чека",
        "receipt_fp": "ФП чека",
        "receipt_fn": "ФН чека"
    }

    # merhe cells with description
    ws.merge_cells('A4:I4')
    ws.merge_cells('J4:L4')
    ws.merge_cells('M4:S4')
    ws.merge_cells('T4:W4')
    ws.merge_cells('X4:AC4')
    ws.merge_cells('AD4:AK4')
    # description cells
    ws['A4'] = 'Общие данные'
    ws['J4'] = 'Данные о сумме операции'
    ws['M4'] = 'Данные контрагента'
    ws['T4'] = 'Детали операции'
    ws['X4'] = 'Данные по карточным операциям'
    ws['AD4'] = 'Дополнительный блок для данных по платежам в бюджет/налогам'

    bold_font = Font(bold=True)
    red_fill = PatternFill(start_color='F77B72',
                           end_color='F77B72', fill_type='solid')
    selected_columns = [column for column in db.TransactionsWithReceipts.__table__.columns.keys(
    ) if column in column_labels]
    headers = [column_labels[column] for column in selected_columns]
    ws.append(headers)

    # write data to file. If has empty cells in 'receipt_fd', 'receipt_fp', 'receipt_fn' draw line in red color
    for row in data:
        row_data = [getattr(row, column) for column in selected_columns]

        if not any(row_data[selected_columns.index(col)] for col in ['receipt_fd', 'receipt_fp', 'receipt_fn']):
            row_to_append = [
                cell if cell is not None else '' for cell in row_data]
            ws.append(row_to_append)
            row_index = ws.max_row
            for r in ws.iter_rows(min_row=row_index, max_row=row_index, min_col=1, max_col=len(row_to_append)):
                for cell in r:
                    cell.fill = red_fill
        else:
            ws.append(row_data)

    # make text bold in description and head
    for cell in ws['A4:AK4'][0]:
        cell.font = bold_font
    for cell in ws['A1:F1'][0]:
        cell.font = bold_font
    for cell in ws['A5:AN5'][0]:
        cell.font = bold_font

    filename = 'exported_data.xlsx'

    # after request delete temp file
    @after_this_request
    def remove_file(response):
        try:
            os.remove(filename)
        except Exception as error:
            app.logger.error(
                "Error removing or closing downloaded file handle", error)
        return response

    wb.save(filename)
    session.close()
    return send_file(filename, as_attachment=True)


def parse_xlsx(file_path, id):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    rows = list(sheet.iter_rows())
    general_headers = rows[0]
    headers = rows[1]

    session = db.SessionLocal()

    for i in range(2, len(rows)):
        r = rows[i]
        item = {
            "account_number": r[0].value,
            "transaction_id": r[1].value,
            "operation_type": r[2].value,
            "operation_category": r[3].value,
            "status": r[4].value,
            "creation_date": str(r[5].value) if r[5].value is not None and str(r[5].value).strip() else None,
            "authorization_date": str(r[6].value) if r[6].value is not None and str(r[6].value).strip() else None,
            "transaction_date": str(r[7].value) if r[7].value is not None and str(r[7].value).strip() else None,
            "original_operation_id": r[8].value,
            "operation_amount_in_currency": r[9].value,
            "operation_currency": r[10].value,
            "amount_in_account_currency": r[11].value,
            "counterparty": r[12].value,
            "counterparty_inn": r[13].value,
            "counterparty_kpp": r[14].value,
            "counterparty_account": r[15].value,
            "counterparty_bank_bik": r[16].value,
            "counterparty_bank_corr_account": r[17].value,
            "counterparty_bank_name": r[18].value,
            "payment_purpose": r[19].value,
            "payment_number": r[20].value,
            "sequence": r[21].value,
            "code": r[22].value,
            "card_number": r[23].value,
            "mcc": r[24].value,
            "place_of_transaction_city": r[25].value,
            "place_of_transaction_country": r[26].value,
            "organization_address": r[27].value,
            "bank": r[28].value,
            "document_creator_status": r[29].value,
            "budget_classification_code": r[30].value,
            "oktmo_code": r[31].value,
            "tax_payment_basis": r[32].value,
            "tax_period_customs_code": r[33].value,
            "tax_document_number": r[34].value,
            "tax_document_date": str(r[35].value),
            "tax_payment_type": r[36].value,
            "komandirovka_id": id
        }

        try:
            transaction = db.Transaction(**item)
            session.add(transaction)
            session.commit()  # Сохраняем каждую транзакцию отдельно
        except DataError as e:
            session.rollback()
            print(f"Ошибка при обработке строки {i + 2}: {e}")
            return {"ok": False, "message": e}
        except Exception as e:
            session.rollback()
            print(f"Ошибка при обработке строки {i + 2}: {e}")
            return {"ok": False, "message": e}

    session.close()   # Закрытие сессии
    return {"ok": True}


@app.route("/add_new_komandirovaniy", methods=['POST'])
def add_new():
    if 'file' not in request.files:
        return 'No file part'
    file = request.files['file']

    if file.filename == '':
        return 'No selected file'

    id = request.form.get('id')

    if file and file.filename.endswith('.xlsx'):
        params = parse_xlsx(file, id)
        return jsonify(params)
    else:
        return {"ok": False, "message": 'Invalid file format'}


@app.route("/qr", methods=['POST'])
def parse_qr():
    if 'file' not in request.files:
        return 'No file part'
    file = request.files['file']

    nparr = np.fromstring(file.read(), np.uint8)
    img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    qreader_out = qreader_reader.detect_and_decode(image=img)

    try:
        result = qreader_out[0]

        params_list = result.split('&')
        params = {}
        for param in params_list:
            key, value = param.split('=')
            params[key.strip()] = value.strip()
        return jsonify(params)
    except:
        return jsonify({"ok": False})


if __name__ == '__main__':
    app.run(debug=True, host="0.0.0.0", port=5000)
